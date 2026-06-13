#!/usr/bin/env python3
"""Extract SAP PO Raw Data rows and Excel fill-scope metadata as JSON."""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import sys
from decimal import Decimal

try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
except ImportError as exc:  # pragma: no cover - runtime guard for UAT hosts.
    print(json.dumps({
        "ok": False,
        "errors": [{"code": "missing_dependency", "message": f"openpyxl is required: {exc}"}],
    }, ensure_ascii=False))
    sys.exit(2)


RAW_DATA_COLUMNS = [
    ("A", "料號", "factory_material_no"),
    ("B", "采購單號", "purchase_order_no"),
    ("C", "採購單狀態", "purchase_order_status"),
    ("D", "采購項次", "purchase_order_line_no"),
    ("E", "法人代碼", "legal_entity_code"),
    ("F", "采購部門代碼", "purchase_department_code"),
    ("G", "采購部門名稱", "purchase_department_name"),
    ("H", "料號", "sap_material_no"),
    ("I", "料號一階類別", "sap_material_lv1"),
    ("J", "料號二階類別", "sap_material_lv2"),
    ("K", "FTV Code", "ftv_code"),
    ("L", "品名 part name", "part_name"),
    ("M", "品名_Detail", "part_name_detail"),
    ("N", "中文譯名", "chinese_translation"),
    ("O", "標準品名", "standard_part_name"),
    ("P", "中文品名_正規化", "normalized_chinese_part_name"),
    ("Q", "正規化", "normalized_item_name"),
    ("R", "品牌", "brand"),
    ("S", "規格 Spec_正規化", "normalized_spec"),
    ("T", "規格 Spec", "spec"),
    ("U", "單位", "uom"),
    ("V", "幣別", "currency"),
    ("W", "采購數量", "purchase_quantity"),
    ("X", "請購單價_USD", "request_unit_price_usd"),
    ("Y", "采購單價_USD", "purchase_unit_price_usd"),
    ("Z", "采購金額_USD", "purchase_amount_usd"),
    ("AA", "請購單價_VND", "request_unit_price_vnd"),
    ("AB", "采購單價_VND", "purchase_unit_price_vnd"),
    ("AC", "采購金額_VND", "purchase_amount_vnd"),
    ("AD", "累計驗收量", "cumulative_accepted_qty"),
    ("AE", "累計領用量", "cumulative_issued_qty"),
    ("AF", "領用金額", "issued_amount"),
    ("AG", "廠商", "vendor_code"),
    ("AH", "廠商簡稱", "vendor_short_name"),
    ("AI", "單價差額", "unit_price_delta"),
    ("AJ", "價差率", "price_delta_rate"),
    ("AK", "采購日期", "purchase_date"),
    ("AL", "預定交期", "scheduled_delivery_date"),
    ("AM", "驗收日期", "acceptance_date"),
    ("AN", "是否逾期", "is_overdue"),
    ("AO", "采購帳號", "buyer_account"),
    ("AP", "采購人員", "buyer_name"),
    ("AQ", "預結報單號", "pre_settlement_no"),
    ("AR", "專案代碼", "project_code"),
    ("AS", "專案", "project_name"),
    ("AT", "MVA or REB", "mva_or_reb"),
    ("AU", "請購單號", "purchase_requisition_no"),
    ("AV", "請購部門", "request_department_code"),
    ("AW", "請購部門名稱", "request_department_name"),
    ("AX", "請購人", "requester_name"),
    ("AY", "請購日期", "request_date"),
    ("AZ", "請購人分機", "requester_extension"),
    ("BA", "請購形式", "request_type"),
    ("BB", "PO release date", "po_release_at"),
    ("BC", "Quotation No", "quotation_no"),
    ("BD", "付款方式", "payment_terms"),
    ("BE", "Delivery term", "delivery_term"),
    ("BF", "稅別/稅率", "tax_rate"),
    ("BG", "費用科目代碼", "expense_account_code"),
    ("BH", "請購單發佈時間", "purchase_requisition_release_time"),
    ("BI", "費用類型", "expense_type"),
    ("BJ", "結報單號", "settlement_no"),
    ("BK", "審核狀態", "review_status"),
    ("BL", "Lv1", "lv1"),
    ("BM", "Lv2", "lv2"),
    ("BN", "Lv3", "lv3"),
]

YELLOW_FILL_ARGB = "FFFFFF00"
BUY_SCOPE_OM = "om_scope"
BUY_SCOPE_MFG_BUY = "mfg_buy"
SCOPE_SOURCE_EXCEL_YELLOW_FILL = "excel_yellow_fill"
SCOPE_SOURCE_DEFAULT_NON_YELLOW = "default_non_yellow"

SUPPLEMENTAL_LV3_PREFIX_BY_PATH = {
    "資訊類::電腦週邊::鍵盤": "ITKEY",
    "資訊類::顯示器::螢幕": "ITMON",
    "資訊類::電腦週邊::滑鼠": "ITMOU",
    "資訊類::電腦::工務電腦": "ITPCA",
    "資訊類::電腦::品保電腦": "ITPCQ",
    "資訊類::電腦::辦公電腦": "ITPCO",
    "資訊類::電腦::工業電腦A級": "ITIPA",
    "資訊類::電腦::工業電腦A+級": "ITIPB",
    "資訊類::電腦::工業電腦A++級": "ITIPC",
    "資訊類::條碼設備::有線標準型條碼掃描器": "ITQWS",
    "資訊類::條碼設備::有線高級型條碼掃描器": "ITQWA",
    "資訊類::條碼設備::無線條碼掃描器": "ITQWL",
    "資訊類::條碼設備::條碼打印機": "ITPRT",
}


def normalize_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value).strip()


def fill_argb(cell) -> str:
    fill = cell.fill
    color = fill.fgColor if fill and fill.fgColor else None
    if not color:
        return ""
    if color.type == "rgb" and color.rgb:
        return str(color.rgb).upper()
    if color.type == "indexed":
        return f"INDEXED:{color.indexed}"
    if color.type == "theme":
        return f"THEME:{color.theme}"
    return ""


def row_fill_scope(ws, row_number: int):
    colors = []
    has_yellow = False
    for excel_column, _header, _field in RAW_DATA_COLUMNS:
        color = fill_argb(ws[f"{excel_column}{row_number}"])
        if color:
            colors.append(color)
        if color == YELLOW_FILL_ARGB:
            has_yellow = True
    return {
        "buy_scope": BUY_SCOPE_OM if has_yellow else BUY_SCOPE_MFG_BUY,
        "scope_source": SCOPE_SOURCE_EXCEL_YELLOW_FILL if has_yellow else SCOPE_SOURCE_DEFAULT_NON_YELLOW,
        "source_fill_color": YELLOW_FILL_ARGB if has_yellow else (colors[0] if colors else ""),
        "has_yellow_fill": has_yellow,
    }


def file_checksum(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_lv_rules(wb):
    if "編碼規則" not in wb.sheetnames:
        return {"lv1_by_name": {}, "lv2_by_lv1_and_name": {}, "warnings": ["Missing 編碼規則 sheet"]}
    ws = wb["編碼規則"]
    lv1_by_name = {}
    lv2_by_lv1_and_name = {}
    for row in ws.iter_rows(min_row=13, values_only=True):
        lv1_code, lv1_name = normalize_text(row[0] if len(row) > 0 else None), normalize_text(row[1] if len(row) > 1 else None)
        lv2_lv1_code = normalize_text(row[3] if len(row) > 3 else None)
        lv2_code = normalize_text(row[4] if len(row) > 4 else None)
        lv2_name = normalize_text(row[5] if len(row) > 5 else None)
        if not any([lv1_code, lv1_name, lv2_lv1_code, lv2_code, lv2_name]):
            break
        if lv1_code and lv1_name:
            lv1_by_name[lv1_name] = lv1_code
        if lv2_lv1_code and lv2_code and lv2_name:
            lv2_by_lv1_and_name[f"{lv2_lv1_code}::{lv2_name}"] = lv2_code
    return {
        "lv1_by_name": lv1_by_name,
        "lv2_by_lv1_and_name": lv2_by_lv1_and_name,
        "supplemental_lv3_prefix_by_path": SUPPLEMENTAL_LV3_PREFIX_BY_PATH,
        "warnings": [],
    }


def expected_prefix(lv1: str, lv2: str, lv_rules, lv3: str = "") -> str:
    supplemental = lv_rules.get("supplemental_lv3_prefix_by_path", {})
    supplemental_prefix = supplemental.get(f"{lv1}::{lv2}::{lv3}")
    if supplemental_prefix:
        return supplemental_prefix
    lv1_code = lv_rules["lv1_by_name"].get(lv1)
    if not lv1_code:
        return "XXXXX" if not lv1 and not lv2 else ""
    lv2_code = lv_rules["lv2_by_lv1_and_name"].get(f"{lv1_code}::{lv2}")
    if not lv2_code:
        return ""
    return f"{lv1_code}{lv2_code}"


def validate_row(row, lv_rules):
    warnings = []
    errors = []
    fields = row["fields"]
    factory_no = fields.get("factory_material_no", "")
    sap_no = fields.get("sap_material_no", "")
    if not factory_no:
        errors.append({"code": "missing_factory_material_no", "message": "Raw Data A factory_material_no is required for import"})
    if factory_no and sap_no and factory_no == sap_no:
        errors.append({"code": "material_no_mixed", "message": "Factory Material No and SAP Material No must not be the same value"})
    prefix = expected_prefix(fields.get("lv1", ""), fields.get("lv2", ""), lv_rules, fields.get("lv3", ""))
    row["expected_factory_prefix"] = prefix
    if fields.get("lv1") or fields.get("lv2"):
        if not prefix:
            warnings.append({
                "code": "lv_rule_missing",
                "message": "LV classification does not have a matching code rule in 編碼規則",
                "lv1": fields.get("lv1", ""),
                "lv2": fields.get("lv2", ""),
            })
        elif factory_no and not factory_no.startswith(prefix):
            warnings.append({
                "code": "factory_prefix_mismatch",
                "message": "Factory Material No prefix does not match 編碼規則",
                "expected_prefix": prefix,
                "factory_material_no": factory_no,
            })
    return errors, warnings


def extract(args):
    workbook_path = os.path.abspath(args.workbook)
    wb = load_workbook(workbook_path, data_only=True, read_only=False)
    if args.sheet not in wb.sheetnames:
        return {"ok": False, "errors": [{"code": "missing_sheet", "message": f"Sheet not found: {args.sheet}"}]}
    ws = wb[args.sheet]
    lv_rules = parse_lv_rules(wb)
    header_errors = []
    for index, (excel_column, expected_header, _field) in enumerate(RAW_DATA_COLUMNS, start=1):
        actual = normalize_text(ws.cell(row=1, column=column_index_from_string(excel_column)).value)
        if actual != expected_header:
            header_errors.append({
                "code": "header_mismatch",
                "excel_column": excel_column,
                "expected": expected_header,
                "actual": actual,
            })
    rows = []
    errors = header_errors[:]
    warnings = [{"code": "lv_rule_warning", "message": warning} for warning in lv_rules.get("warnings", [])]
    for row_number in range(2, ws.max_row + 1):
        scope = row_fill_scope(ws, row_number)
        if args.scope == "yellow-only" and not scope["has_yellow_fill"]:
            continue
        fields = {}
        raw_payload = {}
        is_blank = True
        for excel_column, header, field in RAW_DATA_COLUMNS:
            value = normalize_text(ws[f"{excel_column}{row_number}"].value)
            fields[field] = value
            raw_payload[f"{excel_column}:{header}"] = value
            if value:
                is_blank = False
        if is_blank:
            continue
        row = {
            "source_row_number": row_number,
            **scope,
            "fields": fields,
            "raw_payload_json": raw_payload,
        }
        row_errors, row_warnings = validate_row(row, lv_rules)
        for item in row_errors:
            errors.append({"row": row_number, **item})
        for item in row_warnings:
            warnings.append({"row": row_number, **item})
        rows.append(row)
    scope_counts = {
        BUY_SCOPE_OM: sum(1 for row in rows if row["buy_scope"] == BUY_SCOPE_OM),
        BUY_SCOPE_MFG_BUY: sum(1 for row in rows if row["buy_scope"] == BUY_SCOPE_MFG_BUY),
    }
    return {
        "ok": not header_errors,
        "mode": "preview",
        "scope_mode": args.scope,
        "source_file_name": os.path.basename(workbook_path),
        "workbook_path": workbook_path,
        "source_sheet_name": args.sheet,
        "header_version": "raw-data-a-bn-20260608",
        "source_checksum": file_checksum(workbook_path),
        "total_rows_in_sheet": max(ws.max_row - 1, 0),
        "selected_row_count": len(rows),
        "scope_counts": scope_counts,
        "errors": errors,
        "warnings": warnings,
        "lv_rules": lv_rules,
        "rows": rows,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--sheet", default="Raw Data")
    parser.add_argument("--scope", choices=["yellow-only", "all"], default="yellow-only")
    args = parser.parse_args()
    try:
        result = extract(args)
    except Exception as exc:  # pragma: no cover - CLI safety boundary.
        result = {"ok": False, "errors": [{"code": "extract_failed", "message": str(exc)}]}
    print(json.dumps(result, ensure_ascii=False))
    return 0 if result.get("ok") else 1


if __name__ == "__main__":
    raise SystemExit(main())
